import gc
import io
import re
import time
import zipfile
from datetime import datetime

import pandas as pd
import openpyxl
import msoffcrypto
import streamlit as st

# ============================================================================
# ONE PROD Automation Suite
#   01 — DRR Cleaner
#   02 — ONEPROD Generator
# ============================================================================

st.set_page_config(page_title="ONE PROD Automation Suite", page_icon="🧾", layout="centered")

# Lock Streamlit's own base theme to dark so BaseWeb popovers/menus match the CSS below.
# This runs before the first render, so it applies regardless of the user's system setting.
try:
    from streamlit import config as _stconfig
    if _stconfig.get_option("theme.base") != "dark":
        _stconfig.set_option("theme.base", "dark")
        _stconfig.set_option("theme.primaryColor", "#E0A83C")
        _stconfig.set_option("theme.backgroundColor", "#111726")
        _stconfig.set_option("theme.secondaryBackgroundColor", "#1D2536")
        _stconfig.set_option("theme.textColor", "#E8ECF4")
except Exception:
    pass  # internal API moved — the CSS below still forces the appearance

EXCLUDED_STATUSES = {"LOCKED", "CONFIRMED", "UNLOCKED", "ABORT", "REACTIVE", "BP"}
CLIENT_PRIORITY = ["BPI RECOV 1", "BPI RECOV 2", "BPI RECOV 3", "BPI RECOV PL", "BPI RECOV P", "BPI SPECIAL PROJECT"]

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap');

    /* ------------------------------------------------------------------ */
    /* FIXED THEME — identical on every machine, ignores system light/dark */
    /* ------------------------------------------------------------------ */
    :root {
        --bg-0: #0F1420;      /* page gradient start  */
        --bg-1: #161D2E;      /* page gradient end    */
        --panel: #1D2536;     /* card surface         */
        --panel-2: #232D42;   /* raised surface       */
        --line: #2C3850;      /* borders              */
        --ink: #E8ECF4;       /* main text            */
        --ink-soft: #9AA6BC;  /* secondary text       */
        --gold: #E0A83C;      /* primary accent       */
        --gold-soft: rgba(224,168,60,.14);
        --blue: #5B8DEF;
        --green: #4CC38A;
        --red: #E5645A;
        --mono: 'IBM Plex Mono', ui-monospace, Menlo, monospace;
        --sans: 'Inter', -apple-system, 'Segoe UI', sans-serif;
        --head: 'Space Grotesk', var(--sans);
    }

    .stApp, [data-testid="stAppViewContainer"] {
        background: linear-gradient(160deg, var(--bg-0) 0%, var(--bg-1) 100%) !important;
        color: var(--ink) !important;
        font-family: var(--sans) !important;
    }
    [data-testid="stHeader"] { background: transparent !important; }

    h1, h2, h3, h4 { font-family: var(--head) !important; color: var(--ink) !important; letter-spacing: .2px; }
    h1 { font-size: 1.65rem !important; }
    h2 { font-size: 1.15rem !important; }
    p, li, label, .stMarkdown, [data-testid="stWidgetLabel"] p { color: var(--ink) !important; }
    [data-testid="stCaptionContainer"], .stCaption, small { color: var(--ink-soft) !important; }
    a { color: var(--blue) !important; }
    hr { border-color: var(--line) !important; }

    /* Sidebar */
    [data-testid="stSidebar"] {
        background: #131A2A !important; border-right: 1px solid var(--line) !important;
    }
    [data-testid="stSidebar"] * { color: var(--ink) !important; }

    /* Brand header */
    .brand {
        display:flex; align-items:center; justify-content:space-between;
        border-bottom: 2px solid var(--gold); padding: 4px 0 14px; margin-bottom: 22px;
    }
    .brand .t { font-family: var(--head); font-weight: 700; font-size: 22px; color: var(--ink); }
    .brand .t span { color: var(--gold); }
    .brand .v { font-family: var(--mono); font-size: 11px; color: var(--ink-soft);
                text-transform: uppercase; letter-spacing: 1.6px; }

    /* Step cards */
    .step-card {
        background: var(--panel); border: 1px solid var(--line); border-radius: 14px;
        padding: 18px 20px 6px; margin: 0 0 16px 0;
        box-shadow: 0 2px 10px rgba(0,0,0,.25);
    }
    .step-head { display:flex; align-items:center; gap:12px; margin-bottom: 10px; }
    .step-num {
        width: 30px; height: 30px; border-radius: 50%; flex: 0 0 30px;
        background: var(--gold-soft); border: 1.5px solid var(--gold); color: var(--gold);
        font-family: var(--mono); font-weight: 600; font-size: 14px;
        display:flex; align-items:center; justify-content:center;
    }
    .step-title { font-family: var(--head); font-weight: 600; font-size: 15.5px; color: var(--ink); }
    .step-sub { font-size: 12.5px; color: var(--ink-soft); margin: -6px 0 10px 42px; line-height: 1.55; }

    .rule-box {
        background: var(--panel); border: 1px solid var(--line); border-left: 3px solid var(--gold);
        border-radius: 10px; padding: 14px 18px; margin-bottom: 16px;
        font-size: 13.5px; color: var(--ink-soft) !important; line-height: 1.65;
    }
    .rule-box b { color: var(--ink); }
    .rule-box li { margin-bottom: 5px; color: var(--ink-soft) !important; }
    .rule-box code { background: var(--panel-2); color: var(--gold); padding: 1px 5px; border-radius: 4px; font-family: var(--mono); font-size: 12px; }

    .section-divider { margin: 40px 0 26px; border-top: 1px dashed var(--line); }

    /* File uploader */
    [data-testid="stFileUploader"] section, [data-testid="stFileUploaderDropzone"] {
        background: var(--panel-2) !important; border: 1.5px dashed #46557A !important;
        border-radius: 12px !important; color: var(--ink-soft) !important;
    }
    [data-testid="stFileUploaderDropzone"]:hover { border-color: var(--gold) !important; background: rgba(224,168,60,.06) !important; }
    [data-testid="stFileUploader"] section * { color: var(--ink-soft) !important; }
    [data-testid="stFileUploader"] section button {
        background: transparent !important; border: 1px solid var(--line) !important;
        color: var(--ink) !important; border-radius: 8px !important;
    }
    [data-testid="stFileUploaderFile"] { color: var(--ink) !important; }
    [data-testid="stFileUploaderFile"] * { color: var(--ink-soft) !important; }

    /* Buttons */
    .stButton > button, .stDownloadButton > button {
        border-radius: 10px !important; font-weight: 600 !important; font-family: var(--sans) !important;
        border: 1px solid var(--line) !important; background: var(--panel-2) !important; color: var(--ink) !important;
        transition: transform .05s, filter .15s;
    }
    .stButton > button:hover, .stDownloadButton > button:hover { filter: brightness(1.15); }
    .stButton > button:active { transform: translateY(1px); }
    .stButton > button[kind="primary"], .stDownloadButton > button[kind="primary"] {
        background: linear-gradient(135deg, var(--gold) 0%, #C88F27 100%) !important;
        border: none !important; color: #171207 !important;
    }
    .stButton > button:disabled { opacity: .45 !important; filter: none !important; }

    /* Progress bar */
    .stProgress > div > div > div { background: var(--panel-2) !important; border-radius: 6px; }
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, var(--blue), var(--gold)) !important; border-radius: 6px;
    }
    .stProgress p { color: var(--ink-soft) !important; font-family: var(--mono) !important; font-size: 12px !important; }

    .status-line {
        font-family: var(--mono); font-size: 12.5px; color: var(--ink-soft);
        margin-top: 6px; display: flex; justify-content: space-between;
    }

    /* Metrics */
    [data-testid="stMetric"] {
        background: var(--panel); border: 1px solid var(--line); border-radius: 12px;
        padding: 12px 16px !important;
    }
    [data-testid="stMetricLabel"] p { color: var(--ink-soft) !important; font-family: var(--mono) !important;
        font-size: 11px !important; text-transform: uppercase; letter-spacing: .6px; }
    [data-testid="stMetricValue"] { color: var(--gold) !important; font-family: var(--head) !important; }

    /* Tables & dataframes */
    [data-testid="stTable"] table { background: var(--panel) !important; border-radius: 10px; overflow: hidden; }
    [data-testid="stTable"] th {
        background: var(--panel-2) !important; color: var(--ink-soft) !important;
        font-family: var(--mono) !important; font-size: 11.5px !important;
        text-transform: uppercase; letter-spacing: .5px; border-color: var(--line) !important;
    }
    [data-testid="stTable"] td { color: var(--ink) !important; border-color: var(--line) !important;
        font-size: 13px !important; }

    /* Expanders */
    [data-testid="stExpander"] {
        background: var(--panel) !important; border: 1px solid var(--line) !important; border-radius: 10px !important;
    }
    [data-testid="stExpander"] summary { color: var(--ink) !important; }
    [data-testid="stExpander"] summary:hover { color: var(--gold) !important; }

    /* Alerts (success / info / warning / error) */
    [data-testid="stAlert"] {
        background: var(--panel) !important; border: 1px solid var(--line) !important;
        border-radius: 10px !important; color: var(--ink) !important;
    }
    [data-testid="stAlert"] p { color: var(--ink) !important; }
    div[data-baseweb="notification"] { background: var(--panel) !important; }

    /* Radio / checkbox / multiselect */
    .stRadio label p, .stCheckbox label p { color: var(--ink) !important; }
    .stMultiSelect [data-baseweb="tag"] {
        background: var(--gold-soft) !important; border: 1px solid var(--gold) !important;
    }
    .stMultiSelect [data-baseweb="tag"] span { color: var(--gold) !important; }
    [data-baseweb="select"] > div {
        background: var(--panel-2) !important; border-color: var(--line) !important; color: var(--ink) !important;
    }
    /* dropdown menus render in a portal — style them too */
    [data-baseweb="popover"] [data-baseweb="menu"], ul[data-baseweb="menu"] {
        background: var(--panel-2) !important; border: 1px solid var(--line) !important;
    }
    [data-baseweb="menu"] li, [data-baseweb="menu"] div { color: var(--ink) !important; }
    [data-baseweb="menu"] li:hover { background: var(--gold-soft) !important; }

    /* Spinner text */
    [data-testid="stSpinner"] p { color: var(--ink-soft) !important; }

    /* Result summary card */
    .result-banner {
        background: linear-gradient(135deg, rgba(76,195,138,.12), rgba(91,141,239,.08));
        border: 1px solid rgba(76,195,138,.4); border-radius: 12px;
        padding: 14px 18px; margin: 14px 0; color: var(--ink);
        font-size: 14px;
    }
    .result-banner b { color: var(--green); }
    </style>
    """,
    unsafe_allow_html=True,
)


def step_card_open(num, title, sub=None):
    """Opens a numbered step card. Call step_card_close() after the widgets inside it."""
    st.markdown(
        f'<div class="step-card"><div class="step-head">'
        f'<div class="step-num">{num}</div><div class="step-title">{title}</div></div>'
        + (f'<div class="step-sub">{sub}</div>' if sub else ""),
        unsafe_allow_html=True,
    )


def step_card_close():
    st.markdown("</div>", unsafe_allow_html=True)


# ============================================================================
# SHARED HELPERS
# ============================================================================

def find_header(columns, target):
    target_norm = target.strip().lower()
    for c in columns:
        if str(c).strip().lower() == target_norm:
            return c
    return None


def sanitize_sheet_name(name, used):
    n = str(name or "UNLABELED").strip()
    n = re.sub(r"[\\/\?\*\[\]:]", "-", n)
    n = n[:31] if n else "UNLABELED"
    final, i = n, 2
    while final in used:
        suffix = f" ({i})"
        final = n[: 31 - len(suffix)] + suffix
        i += 1
    used.add(final)
    return final


def sort_client_keys(keys):
    def rank(k):
        try:
            return CLIENT_PRIORITY.index(k.upper())
        except ValueError:
            return len(CLIENT_PRIORITY)
    return sorted(keys, key=lambda k: (rank(k), k))


def format_duration(seconds):
    if seconds is None:
        return "—"
    seconds = max(0, int(round(seconds)))
    if seconds < 60:
        return f"{seconds}s"
    m, s = divmod(seconds, 60)
    return f"{m}m {s:02d}s"


class ProgressTracker:
    """Drives a Streamlit progress bar plus an elapsed-time / ETA status line."""

    def __init__(self, bar, status):
        self.bar = bar
        self.status = status
        self.start = time.time()

    def update(self, pct, label):
        pct = max(0, min(100, pct))
        elapsed = time.time() - self.start
        eta = elapsed * (100 - pct) / pct if pct >= 3 else None
        self.bar.progress(int(pct), text=label)
        eta_text = f"ETA ~{format_duration(eta)}" if eta is not None else "estimating time…"
        self.status.markdown(
            f'<div class="status-line"><span>{label}</span>'
            f'<span>elapsed {format_duration(elapsed)} · {eta_text}</span></div>',
            unsafe_allow_html=True,
        )


# ============================================================================
# AUTOMATION 01 — DRR CLEANER
# ============================================================================

def format_date_value(raw):
    if pd.isna(raw) or raw == "":
        return raw
    s = str(raw).strip()
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$", s)
    if m:
        first, second, year = m.groups()
        return f"{int(second):02d}/{int(first):02d}/{year}"
    return s


def fix_account_no(raw):
    if pd.isna(raw) or raw == "":
        return raw
    s = str(raw).strip()
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    if len(s) == 15:
        return "0" + s
    if len(s) == 8:
        return "000000" + s
    return s


def read_uploaded_table(file_bytes, file_name):
    # Deliberately NOT cached: st.cache_data deep-copies its return value on every access,
    # which for a large DRR means every call silently allocates another full copy in memory.
    # This function is only ever called once per actual read in the pipeline below.
    if file_name.lower().endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, keep_default_na=False)
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, keep_default_na=False)
    return df


def clean_drr(df):
    """Runs the full DRR cleaning pipeline. Returns (groups dict, headers, log lines, stats)."""
    log_lines = []
    headers = list(df.columns)
    date_col = find_header(headers, "Date")
    acct_col = find_header(headers, "Account No.") or find_header(headers, "Account No")
    status_col = find_header(headers, "Status")
    client_col = find_header(headers, "Client")

    log_lines.append(
        f"Columns detected — Date: `{date_col or 'not found'}`, Account No.: `{acct_col or 'not found'}`, "
        f"Status: `{status_col or 'not found'}`, Client: `{client_col or 'not found'}`"
    )

    if date_col:
        df[date_col] = df[date_col].apply(format_date_value)
        log_lines.append("Dates converted to mm/dd/yyyy")
    else:
        log_lines.append("⚠️ No Date column found — skipped")

    if acct_col:
        df[acct_col] = df[acct_col].apply(fix_account_no)
        log_lines.append("Account No. padded (15→16 digits, 8→14 digits)")
    else:
        log_lines.append("⚠️ No Account No. column found — skipped")

    total = len(df)
    if status_col:
        status_upper = df[status_col].astype(str).str.strip().str.upper()
        mask_excluded = status_upper.isin(EXCLUDED_STATUSES)
        excluded_count = int(mask_excluded.sum())
        kept_df = df[~mask_excluded].copy()
        log_lines.append(f"Excluded **{excluded_count:,}** rows (LOCKED / CONFIRMED / UNLOCKED / ABORT / REACTIVE / BP)")
    else:
        kept_df = df.copy()
        excluded_count = 0
        log_lines.append("⚠️ No Status column found — status filter skipped")

    groups = {}
    if client_col:
        kept_df["_client_key"] = kept_df[client_col].astype(str).str.strip().replace("", "UNLABELED")
        for key, sub in kept_df.groupby("_client_key"):
            groups[key] = sub.drop(columns=["_client_key"]).reset_index(drop=True)
        log_lines.append(f"Split into **{len(groups)}** client sheet(s): {', '.join(sort_client_keys(list(groups.keys())))}")
    else:
        groups["ALL"] = kept_df.reset_index(drop=True)
        log_lines.append("⚠️ No Client column found — writing a single sheet")

    stats = {"total": total, "excluded": excluded_count, "kept": len(kept_df)}
    del df, kept_df
    gc.collect()
    return groups, headers, log_lines, stats


def build_drr_workbook(groups, headers):
    output = io.BytesIO()
    used_names = set()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for key in sort_client_keys(list(groups.keys())):
            sheet_df = groups[key][headers]
            sheet_name = sanitize_sheet_name(key, used_names)
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
    return output.getvalue()


def render_drr_cleaner():
    st.header("🧾 Automation 01 — DRR Cleaner")
    st.markdown(
        """
        <div class="rule-box">
        <b>Rules applied to every upload:</b>
        <ul>
            <li><b>Date</b> — reformatted to mm/dd/yyyy</li>
            <li><b>Account No.</b> — 15-digit values padded to 16 (leading 0), 8-digit values padded to 14 (six leading zeros)</li>
            <li><b>Status</b> — rows marked LOCKED, CONFIRMED, UNLOCKED, ABORT, REACTIVE or BP are removed</li>
            <li><b>Client</b> — remaining rows are split into one sheet per client (BPI RECOV 1/2/3, BPI RECOV PL, BPI SPECIAL PROJECT, etc.)</li>
        </ul>
        </div>
        """,
        unsafe_allow_html=True,
    )

    uploaded = st.file_uploader("Upload the DRR file", type=["csv", "xlsx", "xls"], key="drr_upload")

    if uploaded is not None:
        run = st.button("Run cleaning", type="primary", key="drr_run_btn")

        if run:
            with st.spinner("Cleaning..."):
                progress = st.progress(0, text="Reading file...")
                file_bytes = uploaded.getvalue()
                df = read_uploaded_table(file_bytes, uploaded.name)
                progress.progress(30, text="Applying cleaning rules...")
                groups, headers, log_lines, stats = clean_drr(df)
                progress.progress(75, text="Building workbook...")
                workbook_bytes = build_drr_workbook(groups, headers)
                progress.progress(100, text="Done")

            # persist everything needed to re-render after any future rerun
            summary_rows = [{"Sheet": k, "Rows": len(groups[k])} for k in sort_client_keys(list(groups.keys()))]
            st.session_state["drr_result"] = {
                "log_lines": log_lines,
                "stats": stats,
                "summary_rows": summary_rows,
                "workbook_bytes": workbook_bytes,
                "base_name": re.sub(r"\.[^.]+$", "", uploaded.name),
                "groups": groups,   # kept in memory so Automation 02 can use it directly, no re-upload needed
                "headers": headers,
            }

    # Always render from session_state (not just right after clicking Run) so the
    # download button's rerun never wipes the results off the screen.
    result = st.session_state.get("drr_result")
    if result:
        st.success("Cleaning complete.")
        with st.expander("Processing log", expanded=False):
            st.markdown("  \n".join(result["log_lines"]))

        c1, c2, c3 = st.columns(3)
        c1.metric("Rows in file", f"{result['stats']['total']:,}")
        c2.metric("Excluded", f"{result['stats']['excluded']:,}")
        c3.metric("Kept", f"{result['stats']['kept']:,}")

        st.table(pd.DataFrame(result["summary_rows"]))

        st.download_button(
            "⬇️ Download cleaned workbook (.xlsx)",
            data=result["workbook_bytes"],
            file_name=f"{result['base_name']}_CLEANED.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            key="drr_download_btn",
        )
        st.caption("This file stays available above — downloading it again won't clear your results.")


# ============================================================================
# AUTOMATION 02 — ONEPROD GENERATOR
# ============================================================================

def load_oneprod_template(file_bytes):
    """Reads the ONEPROD template: the A:S column names, the U:AU data header block, and
    builds lookup dictionaries from the Reference sheet so every A:S value can be
    pre-computed in Python and saved as a plain value — no formulas, no recalculation
    needed, opens "ready to go" in any spreadsheet app."""
    wb_formulas = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)

    sheet_name = "ONEPROD" if "ONEPROD" in wb_formulas.sheetnames else wb_formulas.sheetnames[0]
    ref_name = "Reference" if "Reference" in wb_formulas.sheetnames else (
        wb_formulas.sheetnames[1] if len(wb_formulas.sheetnames) > 1 else None
    )

    ws = wb_formulas[sheet_name]
    max_col = ws.max_column
    headers_all = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]

    # locate the "S.No" data block
    data_start = None
    for idx, h in enumerate(headers_all, start=1):
        if h is not None and str(h).strip().lower() == "s.no":
            data_start = idx
            break
    if data_start is None:
        raise ValueError("Couldn't find an 'S.No' header in the template — is this the right file?")

    formula_end = data_start - 1
    if formula_end >= 1 and headers_all[formula_end - 1] is None:
        formula_end -= 1  # skip the blank separator column

    data_headers = []
    c = data_start
    while c <= max_col and headers_all[c - 1] is not None:
        data_headers.append(str(headers_all[c - 1]).strip())
        c += 1
    data_end = c - 1

    formula_headers = headers_all[:formula_end]

    status_lookup, reason_lookup = {}, {}
    ref_rows = None
    if ref_name:
        ref_ws_values = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)[ref_name]
        ref_rows = [
            [ref_ws_values.cell(row=r, column=cc).value for cc in range(1, ref_ws_values.max_column + 1)]
            for r in range(1, ref_ws_values.max_row + 1)
        ]
        # Status (col A) -> ONEPROD STATUS (col H); REASON CODE (col J) -> DESCRIPTION (col K)
        for row in ref_rows[1:]:
            if len(row) > 7 and row[0] not in (None, ""):
                status_lookup.setdefault(str(row[0]).strip(), row[7])
            if len(row) > 10 and row[9] not in (None, ""):
                reason_lookup.setdefault(str(row[9]).strip(), row[10])

    return {
        "sheet_name": sheet_name,
        "ref_name": ref_name or "Reference",
        "formula_headers": formula_headers,
        "formula_end": formula_end,
        "data_start": data_start,
        "data_end": data_end,
        "data_headers": data_headers,
        "status_lookup": status_lookup,
        "reason_lookup": reason_lookup,
        "ref_rows": ref_rows,
    }


def _is_zero_like(v):
    return v in ("0.00", "0", 0) or v == 0


def _extract_rfd(remark):
    if not remark:
        return ""
    remark = str(remark)
    idx = remark.find("RFD:")
    if idx == -1:
        return ""
    start = idx + 4
    pipe_idx = remark.find("|", idx)
    if pipe_idx == -1:
        return ""
    return remark[start:pipe_idx].replace("*", "").strip()


def _parse_time(raw):
    if not raw:
        return None
    s = str(raw).strip()
    for fmt in ("%I:%M:%S %p", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return s  # leave unparseable text as-is rather than lose it


def compute_formula_columns_row(date_v, time_w, acct_y, status_ad, remark_ae, ptp_ap, claim_as,
                                 status_lookup, reason_lookup):
    """Reproduces, as plain computed values, exactly what the template's A:S formulas
    would calculate for one row."""
    chosen = claim_as if _is_zero_like(ptp_ap) else ptp_ap
    l_val = None if (chosen in ("", 0, "0.00", None)) else chosen
    reason = _extract_rfd(remark_ae)
    return [
        "BPI CARDS",                                                      # A Financier_Id
        "000" + str(acct_y or ""),                                        # B Application_Id
        acct_y,                                                           # C Customer_Id
        "NDF",                                                           # D User_Id
        date_v,                                                           # E Action_Date
        _parse_time(time_w),                                             # F Action_Time
        status_lookup.get(str(status_ad).strip(), "N/A"),                # G Action_Code
        "NA",                                                            # H Contact_Mode
        "NA",                                                            # I Person_Contacted
        "NA",                                                            # J Place_Contacted
        "PHP",                                                           # K Currency
        l_val,                                                           # L Action_Amount
        date_v,                                                           # M Next_Action_Date
        "00:00",                                                         # N Next_Action_Time
        None,                                                            # O Reminder_Mode
        None,                                                            # P Contacted_By
        remark_ae,                                                        # Q Remarks
        reason if reason else None,                                       # R REASON CODE
        reason_lookup.get(reason.strip(), "N/A") if reason else "N/A",   # S DESCRIPTION
    ]


def generate_oneprod_workbook(template, clients_data: dict, progress_cb=None):
    """clients_data: {client_name: DataFrame (full cleaned columns)}

    Uses openpyxl's write_only mode: rows are streamed straight to the zip on disk instead of
    being kept as Cell objects in memory. Normal-mode openpyxl keeps every cell resident, which
    is what causes MemoryError on large files (hundreds of thousands of rows x dozens of
    columns) — write_only mode discards each row once it's written, so memory stays flat
    regardless of file size.
    """
    wb = openpyxl.Workbook(write_only=True)

    used_names = set()
    data_headers = template["data_headers"]
    formula_headers = template["formula_headers"]
    status_lookup = template["status_lookup"]
    reason_lookup = template["reason_lookup"]
    n_formula_cols = len(formula_headers)

    warnings = []
    n_clients = max(len(clients_data), 1)
    client_names = list(clients_data.keys())

    for i, client_name in enumerate(client_names):
        df = clients_data.pop(client_name)  # free it from the caller's dict as we go
        cols = list(df.columns)
        col_map = {}
        for dh in data_headers:
            match = find_header(cols, dh)
            if match is None:
                warnings.append(f"'{dh}' not found in {client_name} sheet — left blank")
            col_map[dh] = match

        n_rows = len(df)
        sheet_name = sanitize_sheet_name(client_name, used_names)
        ws = wb.create_sheet(title=sheet_name)

        ws.append(list(formula_headers) + list(data_headers))

        # pre-extract data columns as lists for speed
        data_columns = {}
        for dh in data_headers:
            match = col_map[dh]
            data_columns[dh] = df[match].tolist() if match is not None else [""] * n_rows
        del df  # drop the client's cleaned DataFrame now that we only need the extracted lists

        def get(name, r):
            return data_columns[name][r] if name in data_columns else ""

        for r in range(n_rows):
            computed = compute_formula_columns_row(
                date_v=get("Date", r),
                time_w=get("Time", r),
                acct_y=get("Account No.", r),
                status_ad=get("Status", r),
                remark_ae=get("Remark", r),
                ptp_ap=get("PTP Amount", r),
                claim_as=get("Claim Paid Amount", r),
                status_lookup=status_lookup,
                reason_lookup=reason_lookup,
            )
            row_out = computed + [data_columns[dh][r] for dh in data_headers]
            ws.append(row_out)

        del data_columns
        gc.collect()
        if progress_cb:
            progress_cb(int(((i + 1) / n_clients) * 85) + 5, f"Built sheet: {sheet_name}")

    # single shared Reference sheet
    if template["ref_rows"]:
        ref_ws = wb.create_sheet(title=template["ref_name"])
        for row_vals in template["ref_rows"]:
            ref_ws.append(row_vals)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), warnings


def render_oneprod_generator():
    st.header("📊 Automation 02 — ONEPROD Generator")
    st.markdown(
        """
        <div class="rule-box">
        Takes a cleaned client sheet (from Automation 01 above, or an uploaded cleaned workbook) and drops it
        into the ONEPROD template: <b>U:AU</b> is filled with the client's data — matched by header name, not position —
        and <b>A:S</b> is pre-computed in Python using the same logic as the template's formulas (including the
        Reference-sheet lookups) and saved as plain values. No formulas, nothing to recalculate — the file opens
        with every column already filled in.
        </div>
        """,
        unsafe_allow_html=True,
    )

    template_file = st.file_uploader(
        "Upload the ONEPROD template (contains the ONEPROD formulas + the Reference sheet)",
        type=["xlsx"], key="tmpl_upload",
    )

    drr_result = st.session_state.get("drr_result")
    source_mode = "upload"
    if drr_result:
        source_mode = st.radio(
            "Which cleaned DRR data should feed the ONEPROD sheets?",
            ["Use the file I just cleaned above", "Upload a different cleaned workbook"],
            key="oneprod_source_mode",
        )
        source_mode = "session" if source_mode.startswith("Use the file") else "upload"

    cleaned_groups = None
    if source_mode == "session" and drr_result:
        cleaned_groups = drr_result["groups"]
    else:
        cleaned_upload = st.file_uploader(
            "Upload the cleaned DRR workbook (output of Automation 01)", type=["xlsx"], key="cleaned_upload"
        )
        if cleaned_upload is not None:
            xls = pd.ExcelFile(io.BytesIO(cleaned_upload.getvalue()))
            cleaned_groups = {
                sn: xls.parse(sn, dtype=str, keep_default_na=False) for sn in xls.sheet_names
            }

    if template_file is None or cleaned_groups is None:
        st.info("Upload the ONEPROD template and provide cleaned DRR data to continue.")
        return

    client_options = sort_client_keys(list(cleaned_groups.keys()))
    selected_clients = st.multiselect(
        "Generate ONEPROD sheet(s) for:", client_options, default=client_options, key="client_select"
    )

    run = st.button("Generate ONEPROD workbook", type="primary", key="oneprod_run_btn", disabled=not selected_clients)

    if run:
        with st.spinner("Generating..."):
            progress = st.progress(0, text="Reading template...")
            template = load_oneprod_template(template_file.getvalue())
            progress.progress(5, text="Building sheets...")

            def cb(pct, label):
                progress.progress(min(pct, 99), text=label)

            clients_data = {c: cleaned_groups[c] for c in selected_clients}
            workbook_bytes, warnings = generate_oneprod_workbook(template, clients_data, progress_cb=cb)
            progress.progress(98, text="Indexing sheets...")

            # keep per-client DataFrames in memory so Automation 03 doesn't need a re-upload
            xls = pd.ExcelFile(io.BytesIO(workbook_bytes))
            client_frames = {
                sn: xls.parse(sn, dtype=str, keep_default_na=False)
                for sn in xls.sheet_names if sn != template["ref_name"]
            }
            progress.progress(100, text="Done")

        st.session_state["oneprod_result"] = {
            "workbook_bytes": workbook_bytes,
            "warnings": warnings,
            "selected_clients": selected_clients,
            "client_frames": client_frames,
        }

    result = st.session_state.get("oneprod_result")
    if result:
        st.success("ONEPROD workbook generated.")
        if result["warnings"]:
            with st.expander("⚠️ Warnings", expanded=False):
                for w in result["warnings"]:
                    st.write("- " + w)
        st.write(f"Sheets generated: {', '.join(result['selected_clients'])} + Reference")

        st.download_button(
            "⬇️ Download ONEPROD workbook (.xlsx)",
            data=result["workbook_bytes"],
            file_name="ONEPROD_OUTPUT.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            key="oneprod_download_btn",
        )
        st.caption("This file stays available above — downloading it again won't clear your results.")


# ============================================================================
# AUTOMATION 03 — FINAL EXPORT (per-client file, columns A:Q, password-protected)
# ============================================================================

FINAL_COLUMNS = [
    "Financier_Id", "Application_Id", "Customer_Id", "User_Id", "Action_Date", "Action_Time",
    "Action_Code", "Contact_Mode", "Person_Contacted", "Place_Contacted", "Currency",
    "Action_Amount", "Next_Action_Date", "Next_Action_Time", "Reminder_Mode", "Contacted_By", "Remarks",
]

CLIENT_FILE_CODES = {
    "BPI RECOV 1": "Reco1",
    "BPI RECOV 2": "Reco2",
    "BPI RECOV 3": "Reco3",
    "BPI RECOV PL": "RecoPL",
    "BPI SPECIAL PROJECT": "Revival",
}


def quarter_password(month, year):
    if month in (1, 2, 3):
        q = 1
    elif month in (4, 5, 6):
        q = 2
    elif month in (7, 8, 9):
        q = 3
    else:
        q = 4
    return f"MAD_{q}Q{year}"


def mmddyyyy_to_yyyymmdd(date_str):
    if not date_str:
        return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", str(date_str).strip())
    if not m:
        return None
    mm, dd, yyyy = m.groups()
    return f"{yyyy}{int(mm):02d}{int(dd):02d}"


def split_client_by_date(df):
    """Splits one client's DataFrame into {mm/dd/yyyy: sub-DataFrame}, preserving row order
    within each date. Falls back to a single 'UNKNOWNDATE' group if no date column is found."""
    cols = list(df.columns)
    date_col = find_header(cols, "Date") or find_header(cols, "Action_Date")
    if date_col is None:
        return {"UNKNOWNDATE": df}
    keys = df[date_col].astype(str).str.strip()
    groups = {}
    for key, sub in df.groupby(keys, sort=False):
        groups[key if key else "UNKNOWNDATE"] = sub.reset_index(drop=True)
    return groups


def date_folder_from_filename(filename):
    m = re.match(r"^CC_(\d{8})_", filename)
    return m.group(1) if m else "UNKNOWNDATE"


def build_final_client_file(df, client_name, user_id_lookup=None):
    """df: DataFrame with the ONEPROD sheet's full column set (A:AU) for one client.
    user_id_lookup: optional dict mapping Customer_Id -> Unit Code. If given, overrides the
    User_Id column (falls back to 'N/A' per row when a Customer_Id isn't found — using 'N/A' rather
    than Excel's literal '#N/A' so it survives round-trips through pandas as plain text, matching what
    an Excel VLOOKUP would show). If None, User_Id keeps whatever value it already had (e.g. 'NDF')."""
    cols = list(df.columns)
    col_map = {h: find_header(cols, h) for h in FINAL_COLUMNS}
    missing = [h for h, m in col_map.items() if m is None]

    n_rows = len(df)
    sub_columns = {
        h: (df[col_map[h]].tolist() if col_map[h] is not None else [""] * n_rows)
        for h in FINAL_COLUMNS
    }

    if user_id_lookup is not None:
        customer_ids = sub_columns.get("Customer_Id", [""] * n_rows)
        sub_columns["User_Id"] = [
            user_id_lookup.get(str(cid).strip(), "N/A") if cid not in (None, "") else "N/A"
            for cid in customer_ids
        ]

    date_col_vals = sub_columns.get("Action_Date", [])
    first_date = next((v for v in date_col_vals if v not in (None, "")), None)
    yyyymmdd = mmddyyyy_to_yyyymmdd(first_date)

    code = CLIENT_FILE_CODES.get(client_name.strip().upper())
    code_warning = None
    if code is None:
        code = re.sub(r"[^A-Za-z0-9]", "", client_name)
        code_warning = f"'{client_name}' has no known file-name code — using '{code}'. Let me know the right one."

    date_warning = None
    if yyyymmdd is None:
        yyyymmdd = "UNKNOWNDATE"
        date_warning = f"Couldn't read a valid date for {client_name} — filename uses 'UNKNOWNDATE'."

    filename = f"CC_{yyyymmdd}_{code}_Madrid.xlsx"

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append([h.lower() for h in FINAL_COLUMNS])
    n_cols = len(FINAL_COLUMNS)
    for r in range(n_rows):
        ws.append([sub_columns[h][r] for h in FINAL_COLUMNS])
    del sub_columns

    # Leave one blank row between the data and EOF, then restrict AutoFilter to header+data
    # only. Excel auto-detects a filter range as the contiguous block starting at the header;
    # the blank row breaks that contiguity, so filtering/sorting the data can never touch EOF.
    last_data_row = n_rows + 1
    ws.append([None] * n_cols)          # blank separator row
    ws.append(["EOF"] * n_cols)         # EOF marker row
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(n_cols)}{last_data_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if yyyymmdd != "UNKNOWNDATE":
        pw_year, pw_month = int(yyyymmdd[:4]), int(yyyymmdd[4:6])
    else:
        now = datetime.now()
        pw_year, pw_month = now.year, now.month
    password = quarter_password(pw_month, pw_year)

    office_file = msoffcrypto.OfficeFile(buf)
    encrypted = io.BytesIO()
    office_file.encrypt(password, encrypted)

    warnings = [w for w in (missing and f"{client_name}: missing columns {missing}", code_warning, date_warning) if w]
    return filename, encrypted.getvalue(), password, warnings


def parse_user_id_reference(file_bytes):
    """Reference workbook: column A = Customer No. (account no., same padding as Customer_Id),
    column B = Unit Code (the value to place in User_Id). Returns a dict, first match wins."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    lookup = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        if key in (None, ""):
            continue
        key = str(key).strip()
        lookup.setdefault(key, val)
    return lookup


def render_final_export():
    st.header("🔐 Automation 03 — Final Export")
    st.markdown(
        """
        <div class="rule-box">
        Takes each client's ONEPROD sheet (from Automation 02 above, or an uploaded ONEPROD workbook) and, for every
        client, produces a separate file containing only <b>columns A–Q</b> (Financier_Id through Remarks, lowercased),
        followed by a trailing <b>EOF</b> row. Each file is named
        <code>CC_yyyymmdd_&lt;code&gt;_Madrid.xlsx</code> — the date comes from that client's Action_Date column — and is
        password-protected using the quarter the report date falls in
        (Q1 Jan–Mar, Q2 Apr–Jun, Q3 Jul–Sep, Q4 Oct–Dec → <code>MAD_&lt;Q&gt;Q&lt;year&gt;</code>).
        All files are bundled into one .zip to download.
        </div>
        """,
        unsafe_allow_html=True,
    )

    oneprod_result = st.session_state.get("oneprod_result")
    source_mode = "upload"
    if oneprod_result:
        source_mode = st.radio(
            "Which ONEPROD data should this export from?",
            ["Use the ONEPROD workbook generated above", "Upload a different ONEPROD workbook"],
            key="final_source_mode",
        )
        source_mode = "session" if source_mode.startswith("Use the ONEPROD") else "upload"

    client_frames = None
    if source_mode == "session" and oneprod_result:
        client_frames = oneprod_result["client_frames"]
    else:
        oneprod_upload = st.file_uploader(
            "Upload the ONEPROD workbook (output of Automation 02)", type=["xlsx"], key="final_upload"
        )
        if oneprod_upload is not None:
            xls = pd.ExcelFile(io.BytesIO(oneprod_upload.getvalue()))
            client_frames = {
                sn: xls.parse(sn, dtype=str, keep_default_na=False)
                for sn in xls.sheet_names if sn.strip().lower() != "reference"
            }

    if client_frames is None:
        st.info("Generate or upload an ONEPROD workbook to continue.")
        return

    client_options = sort_client_keys(list(client_frames.keys()))
    selected_clients = st.multiselect(
        "Export final file(s) for:", client_options, default=client_options, key="final_client_select"
    )

    run = st.button("Generate final export", type="primary", key="final_run_btn", disabled=not selected_clients)

    if run:
        with st.spinner("Building final files..."):
            progress = st.progress(0, text="Starting...")
            files = []
            all_warnings = []
            n = max(len(selected_clients), 1)
            for i, client_name in enumerate(selected_clients):
                date_groups = split_client_by_date(client_frames[client_name])
                for date_str, sub_df in date_groups.items():
                    filename, enc_bytes, password, warnings = build_final_client_file(
                        sub_df, client_name
                    )
                    files.append({"filename": filename, "bytes": enc_bytes, "password": password, "client": client_name})
                    all_warnings.extend(warnings)
                progress.progress(int(((i + 1) / n) * 90), text=f"Built {client_name} ({len(date_groups)} date(s))")

            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for f in files:
                    folder = date_folder_from_filename(f["filename"])
                    zf.writestr(f"{folder}/{f['filename']}", f["bytes"])
            progress.progress(100, text="Done")

        st.session_state["final_result"] = {
            "files": files,
            "warnings": all_warnings,
            "zip_bytes": zip_buf.getvalue(),
        }

    result = st.session_state.get("final_result")
    if result:
        st.success(f"{len(result['files'])} file(s) ready.")
        if result["warnings"]:
            with st.expander("⚠️ Warnings", expanded=False):
                for w in result["warnings"]:
                    st.write("- " + w)

        rows = [{"File": f["filename"], "Password": f["password"]} for f in result["files"]]
        st.table(pd.DataFrame(rows))

        st.download_button(
            "⬇️ Download all final files (.zip)",
            data=result["zip_bytes"],
            file_name="ONEPROD_FINAL_EXPORT.zip",
            mime="application/zip",
            type="primary",
            key="final_download_btn",
        )
        st.caption("This file stays available above — downloading it again won't clear your results.")


# ============================================================================
# AUTOMATION 04 — USER ID LOOKUP + FINAL PACKAGE
# ============================================================================

def render_user_id_lookup_export():
    st.header("🔎 Automation 04 — User ID Lookup")
    st.markdown(
        """
        <div class="rule-box">
        Same final export as Automation 03, plus one optional step: upload a <b>User ID reference</b> file
        (column A = Customer No., column B = Unit Code) and the <b>user_id</b> column gets filled in by matching
        each row's Customer_Id against it — the same result as the <code>VLOOKUP(customer_id, ref, 2, FALSE)</code>
        formula. <b>The upload is optional</b> — skip it and every file is still generated normally, just with
        <code>user_id</code> left as-is. The trailing <b>EOF</b> row is also fixed here: a blank row now sits
        between the data and EOF, and AutoFilter is scoped to header+data only, so filtering or sorting the data
        can never hide or move the EOF row.
        </div>
        """,
        unsafe_allow_html=True,
    )

    oneprod_result = st.session_state.get("oneprod_result")
    source_mode = "upload"
    if oneprod_result:
        source_mode = st.radio(
            "Which ONEPROD data should this export from?",
            ["Use the ONEPROD workbook generated above", "Upload a different ONEPROD workbook"],
            key="v4_source_mode",
        )
        source_mode = "session" if source_mode.startswith("Use the ONEPROD") else "upload"

    client_frames = None
    if source_mode == "session" and oneprod_result:
        client_frames = oneprod_result["client_frames"]
    else:
        oneprod_upload = st.file_uploader(
            "Upload the ONEPROD workbook (output of Automation 02)", type=["xlsx"], key="v4_upload"
        )
        if oneprod_upload is not None:
            xls = pd.ExcelFile(io.BytesIO(oneprod_upload.getvalue()))
            client_frames = {
                sn: xls.parse(sn, dtype=str, keep_default_na=False)
                for sn in xls.sheet_names if sn.strip().lower() != "reference"
            }

    if client_frames is None:
        st.info("Generate or upload an ONEPROD workbook to continue.")
        return

    ref_upload = st.file_uploader(
        "User ID reference file (optional)", type=["xlsx"], key="v4_ref_upload"
    )

    client_options = sort_client_keys(list(client_frames.keys()))
    selected_clients = st.multiselect(
        "Export final file(s) for:", client_options, default=client_options, key="v4_client_select"
    )

    run = st.button("Generate final package", type="primary", key="v4_run_btn", disabled=not selected_clients)

    if run:
        with st.spinner("Building final files..."):
            progress = st.progress(0, text="Starting...")

            user_id_lookup = None
            if ref_upload is not None:
                user_id_lookup = parse_user_id_reference(ref_upload.getvalue())
                progress.progress(10, text=f"Loaded {len(user_id_lookup):,} User ID entries")

            files = []
            all_warnings = []
            n = max(len(selected_clients), 1)
            for i, client_name in enumerate(selected_clients):
                date_groups = split_client_by_date(client_frames[client_name])
                for date_str, sub_df in date_groups.items():
                    filename, enc_bytes, password, warnings = build_final_client_file(
                        sub_df, client_name, user_id_lookup=user_id_lookup
                    )
                    files.append({"filename": filename, "bytes": enc_bytes, "password": password, "client": client_name})
                    all_warnings.extend(warnings)
                progress.progress(10 + int(((i + 1) / n) * 85), text=f"Built {client_name} ({len(date_groups)} date(s))")

            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for f in files:
                    folder = date_folder_from_filename(f["filename"])
                    zf.writestr(f"{folder}/{f['filename']}", f["bytes"])
            progress.progress(100, text="Done")

        st.session_state["v4_result"] = {
            "files": files,
            "warnings": all_warnings,
            "zip_bytes": zip_buf.getvalue(),
            "used_lookup": ref_upload is not None,
        }

    result = st.session_state.get("v4_result")
    if result:
        st.success(
            f"{len(result['files'])} file(s) ready"
            + (" — User ID lookup applied." if result["used_lookup"] else " — no reference uploaded, user_id left as-is.")
        )
        if result["warnings"]:
            with st.expander("⚠️ Warnings", expanded=False):
                for w in result["warnings"]:
                    st.write("- " + w)

        rows = [{"File": f["filename"], "Password": f["password"]} for f in result["files"]]
        st.table(pd.DataFrame(rows))

        st.download_button(
            "⬇️ Download all final files (.zip)",
            data=result["zip_bytes"],
            file_name="ONEPROD_FINAL_EXPORT_V4.zip",
            mime="application/zip",
            type="primary",
            key="v4_download_btn",
        )
        st.caption("This file stays available above — downloading it again won't clear your results.")


# ============================================================================
# UNIFIED PIPELINE — everything in one run
# ============================================================================

PIPELINE_STEPS = [
    "Clean DRR", "Load template", "Build ONEPROD sheets",
    "User ID lookup", "Build final files", "Package zip",
]


def run_full_pipeline(drr_bytes, drr_name, template_bytes, ref_bytes, selected_clients, tracker,
                       include_audit_workbook=True):
    tracker.update(2, "Reading DRR file...")
    df = read_uploaded_table(drr_bytes, drr_name)
    tracker.update(8, f"Loaded {len(df):,} rows — cleaning...")

    groups, headers, log_lines, stats = clean_drr(df)
    del df
    gc.collect()
    groups = {k: v for k, v in groups.items() if k in selected_clients}
    # capture row counts now — generate_oneprod_workbook drains `groups` as it processes clients
    summary_rows = [{"Sheet": k, "Rows": len(groups[k])} for k in sort_client_keys(list(groups.keys()))]
    tracker.update(25, "DRR cleaned & split by client")

    template = load_oneprod_template(template_bytes)
    tracker.update(30, "ONEPROD template loaded")

    def oneprod_cb(pct, label):
        tracker.update(30 + (pct / 100) * 30, label)

    workbook_bytes, oneprod_warnings = generate_oneprod_workbook(template, groups, progress_cb=oneprod_cb)
    del groups
    gc.collect()
    tracker.update(60, "ONEPROD sheets built")

    xls = pd.ExcelFile(io.BytesIO(workbook_bytes))
    client_sheet_names = [sn for sn in xls.sheet_names if sn != template["ref_name"]]
    tracker.update(63, "Preparing final export...")

    user_id_lookup = None
    if ref_bytes:
        user_id_lookup = parse_user_id_reference(ref_bytes)
        tracker.update(67, f"Loaded {len(user_id_lookup):,} User ID entries")

    files = []
    all_warnings = list(oneprod_warnings)
    n = max(len(client_sheet_names), 1)
    for i, client_name in enumerate(client_sheet_names):
        cdf = xls.parse(client_name, dtype=str, keep_default_na=False)  # read one sheet at a time
        date_groups = split_client_by_date(cdf)
        del cdf
        for date_str, sub_df in date_groups.items():
            filename, enc_bytes, password, warns = build_final_client_file(
                sub_df, client_name, user_id_lookup=user_id_lookup
            )
            files.append({"filename": filename, "bytes": enc_bytes, "password": password,
                           "client": client_name, "rows": len(sub_df)})
            all_warnings.extend(warns)
        del date_groups
        gc.collect()
        tracker.update(67 + ((i + 1) / n) * 27, f"Finalized {client_name}")

    tracker.update(96, "Packaging zip...")
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in files:
            folder = date_folder_from_filename(f["filename"])
            zf.writestr(f"{folder}/{f['filename']}", f["bytes"])
        if include_audit_workbook:
            zf.writestr("ONEPROD_WORKBOOK.xlsx", workbook_bytes)
    tracker.update(100, "Done")

    return {
        "files": files,
        "warnings": all_warnings,
        "zip_bytes": zip_buf.getvalue(),
        "stats": stats,
        "summary_rows": summary_rows,
        "used_lookup": ref_bytes is not None,
        "log_lines": log_lines,
    }


def render_unified_pipeline():
    st.markdown(
        '<div class="brand"><div class="t">ONE PROD <span>AUTOMATION</span></div>'
        '<div class="v">MADRECO · one-click pipeline</div></div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="rule-box">Upload the files below, press <b>Run</b>, download the results. '
        'The pipeline cleans the DRR, builds the ONEPROD sheets, applies the User ID lookup, and exports '
        'one password-protected <code>CC_yyyymmdd_&lt;client&gt;_Madrid.xlsx</code> per client per day — '
        'all bundled in a single zip organized by date.</div>',
        unsafe_allow_html=True,
    )

    # ---------------- STEP 1 — files ----------------
    step_card_open(1, "Upload your files", "The first two are required. The User ID reference is optional — skip it and user_id stays as-is.")
    c1, c2 = st.columns(2)
    with c1:
        drr_file = st.file_uploader("DRR file (.csv / .xlsx)", type=["csv", "xlsx", "xls"], key="u_drr")
    with c2:
        template_file = st.file_uploader("ONEPROD template (.xlsx)", type=["xlsx"], key="u_template")
    ref_file = st.file_uploader("User ID reference — optional (.xlsx)", type=["xlsx"], key="u_ref")
    step_card_close()

    # ---------------- STEP 2 — options ----------------
    step_card_open(2, "Choose what to process", "Clients are detected automatically from the DRR's Client column.")
    selected_clients = None
    all_clients_found = None
    if drr_file is not None:
        preview_df = read_uploaded_table(drr_file.getvalue(), drr_file.name)
        client_col = find_header(list(preview_df.columns), "Client")
        if client_col:
            all_clients_found = sort_client_keys(
                [c.strip() for c in preview_df[client_col].astype(str).str.strip().unique() if c.strip()]
            )
            selected_clients = st.multiselect(
                "Clients to process", all_clients_found, default=all_clients_found, key="u_client_select"
            )
        del preview_df
        gc.collect()
    else:
        st.caption("⬆ Upload a DRR file first — the client list will appear here.")

    include_audit = st.checkbox(
        "Include the full ONEPROD workbook (A:AU, all clients) in the download for audit purposes",
        value=True, key="u_include_audit",
        help="Turn this off for very large DRRs to cut peak memory usage — you'll still get every "
             "password-protected CC_*.xlsx file, just not the combined audit copy.",
    )
    step_card_close()

    # ---------------- STEP 3 — run ----------------
    step_card_open(3, "Run", "Progress, elapsed time and ETA will show below while it works.")
    ready = drr_file is not None and template_file is not None and (selected_clients is None or selected_clients)
    if not ready:
        missing = []
        if drr_file is None:
            missing.append("DRR file")
        if template_file is None:
            missing.append("ONEPROD template")
        if drr_file is not None and template_file is not None:
            missing.append("at least one client selected")
        st.caption("Waiting for: " + ", ".join(missing))
    run = st.button("▶  Run full automation", type="primary", key="u_run_btn", disabled=not ready,
                    use_container_width=True)
    progress_slot = st.container()
    step_card_close()

    if run:
        with progress_slot:
            bar = st.progress(0, text="Starting...")
            status = st.empty()
        tracker = ProgressTracker(bar, status)

        drr_bytes = drr_file.getvalue()
        drr_name = drr_file.name
        template_bytes = template_file.getvalue()
        ref_bytes = ref_file.getvalue() if ref_file is not None else None
        use_clients = (selected_clients if selected_clients else all_clients_found) or ["ALL"]

        with st.spinner("Running the full pipeline..."):
            result = run_full_pipeline(
                drr_bytes, drr_name, template_bytes, ref_bytes, use_clients, tracker,
                include_audit_workbook=include_audit,
            )

        st.session_state["pipeline_result"] = result

    # ---------------- Results ----------------
    result = st.session_state.get("pipeline_result")
    if result:
        st.markdown(
            f'<div class="result-banner">✅ <b>Done</b> — {len(result["files"])} file(s) ready'
            + (" · User ID lookup applied." if result["used_lookup"] else " · no User ID reference uploaded.")
            + "</div>",
            unsafe_allow_html=True,
        )

        m1, m2, m3 = st.columns(3)
        m1.metric("Rows in DRR", f"{result['stats']['total']:,}")
        m2.metric("Excluded", f"{result['stats']['excluded']:,}")
        m3.metric("Kept", f"{result['stats']['kept']:,}")

        if result["warnings"]:
            with st.expander(f"⚠️ Warnings ({len(result['warnings'])})", expanded=False):
                for w in result["warnings"]:
                    st.write("- " + w)

        with st.expander("Processing log", expanded=False):
            st.markdown("  \n".join(result["log_lines"]))

        rows = [{"File": f["filename"], "Rows": f["rows"], "Password": f["password"]} for f in result["files"]]
        st.table(pd.DataFrame(rows))

        st.download_button(
            "⬇️  Download everything (.zip)",
            data=result["zip_bytes"],
            file_name="ONEPROD_FINAL_EXPORT.zip",
            mime="application/zip",
            type="primary",
            key="u_download_btn",
            use_container_width=True,
        )
        st.caption(
            "The .zip includes each password-protected CC_*.xlsx (organized in per-date folders) plus "
            "ONEPROD_WORKBOOK.xlsx for audit. This stays available above — downloading again won't clear your results."
        )


# ============================================================================
# PAGE LAYOUT
# ============================================================================

with st.sidebar:
    st.markdown(
        '<div style="font-family:var(--head);font-weight:700;font-size:17px;padding:4px 0 2px;">'
        '🧾 ONE PROD</div>'
        '<div style="font-family:var(--mono);font-size:10.5px;color:var(--ink-soft);'
        'text-transform:uppercase;letter-spacing:1.4px;margin-bottom:14px;">Automation Suite</div>',
        unsafe_allow_html=True,
    )
    mode = st.radio(
        "Mode", ["All-in-one (recommended)", "Step-by-step (advanced)"], key="app_mode",
        help="All-in-one runs the whole pipeline with one click. Step-by-step lets you run each stage "
             "separately and download intermediate files.",
    )
    st.markdown('<div class="section-divider" style="margin:18px 0;"></div>', unsafe_allow_html=True)
    st.caption(
        "**Quick guide**  \n"
        "1. Upload DRR + ONEPROD template  \n"
        "2. (Optional) User ID reference  \n"
        "3. Press Run, then download the zip  \n\n"
        "File passwords follow the quarter of each file's report date "
        "(e.g. Jul–Sep 2026 → `MAD_3Q2026`)."
    )

if mode.startswith("All-in-one"):
    render_unified_pipeline()
else:
    st.markdown(
        '<div class="brand"><div class="t">ONE PROD <span>AUTOMATION</span></div>'
        '<div class="v">Step-by-step mode</div></div>',
        unsafe_allow_html=True,
    )
    render_drr_cleaner()
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    render_oneprod_generator()
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    render_final_export()
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    render_user_id_lookup_export()